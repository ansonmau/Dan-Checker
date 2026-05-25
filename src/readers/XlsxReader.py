import openpyxl 

class XlsxReader:                                                                                                                                                                  
    def __init__(self):                                                                                                                                                            
        self._file_path = None                                                                                                                                                     
        self._workbook = None                                                                                                                                                      
        self._sheet = None
        self._headers = {}
                                                                                                                                                                             
    def set_target_file(self, file_path: str):
        self._file_path = file_path                                                                                                                                                
        self._workbook = None
        self._sheet = None

    def open(self, sheet_name: str = ""):                                                                                                                                        
        if not self._file_path:
          raise ValueError("No target file set. Call set_target_file() first.")                                                                                                  
        self._workbook = openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
        self._sheet = self._workbook[sheet_name] if sheet_name else self._workbook.active                                                                                          
        self._get_headers()
        return 0
                                                                                                                                                                             
    def get_value(self, row: int, col: int):                                                                                                                                       
        self._check_sheet_loaded()
        return self._sheet.cell(row=row, column=col).value
    
    def get_num_rows(self):
        self._check_sheet_loaded()
        return self._sheet.max_row

    def get_num_cols(self):
        self._check_sheet_loaded()
        return self._sheet.max_column

    def header(self, value):
        self._check_sheet_loaded()
        return self._headers[value]

    def _get_headers(self):
        self._check_sheet_loaded()
        for col in range(1, self.get_num_cols() + 1):
            val = self.get_value(row=1, col=col)
            if (val and str(val).strip() != ""):
                self._headers.update({f"{val}": col})

    def _check_sheet_loaded(self):
        if self._sheet is None:                                                                                                                                                    
          raise RuntimeError("File not opened. Call open() first.")
